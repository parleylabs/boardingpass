from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from api import settings
from api.models.device_key import DeviceKey

class Command(BaseCommand):

    def handle(self, *args, **options):
        print("Running Test Code")
        path = str(settings.BASE_DIR) +'/api/management/commands/api_devicekey.xlsx'
        wb2 = load_workbook(path)
        sheet = wb2.active
        m_row = sheet.max_row
        m_col = sheet.max_column

        index_keys = {}

        # Get headers of each column
        for i in range(1, 2):
            for j in range(1, m_col + 1):
                field_value = sheet.cell(row=i, column=j).value
                if field_value:
                    index_keys[j] = field_value

        for i in range(2, m_row + 1):
            print()
            print(f'processing row {i}')
            data_object = {}

            for j in range(1, m_col + 1):
                label = None
                field_value = sheet.cell(row=i, column=j).value
                if j in index_keys:
                    label = index_keys[j]

                if label:
                    data_object[label] = str(field_value)

            DeviceKey.objects.create(
                dev_eui = data_object['dev_eui'],
                app_eui = data_object['app_eui'],
                app_key = data_object['app_key'],
                claimed_status = data_object['claimed_status'],
                organization_id = data_object['organization_id'],
            )



